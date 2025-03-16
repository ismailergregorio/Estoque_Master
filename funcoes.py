from funcoes import*
import tkinter as tk
from tkinter import messagebox
from datetime import datetime
import openpyxl
import os
from openpyxl.utils.datetime import from_excel
from openpyxl import load_workbook
from tkinter import *
from tkinter import ttk
from tkcalendar import DateEntry

cadastro = "Base de dados\CADASTRO.xlsx"
dados_de_saida = "Base de dados\saida.xlsx"
dados_de_entrada = "Base de dados\entrada.xlsx"
dados_de_estoque = "Base de dados\Estoque.xlsx"
dados_de_cadastro_nf = "Base de dados\Cadastro NF.xlsx"
dados_de_cadastro_fornecedor = "Base de dados\Cadastro Fornecedor.xlsx"

nomes_dos_setores = ["RH", "ESCRITORIO", "SEGURANÇA", "COSINHA LIMPESA", "COSINHA COMIDA", "TRADE", "OPERACIONAL"]

def mostrar_alerta(mesagem):
    # Mostra uma caixa de alerta com título e mensagem
    if messagebox.askyesno("Atenção", mesagem):
        return True
    else:
        return False

def alerta_erro(mensagem, titulo="Erro"):
    """
    Exibe um alerta de erro em uma janela pop-up.

    Args:
        mensagem (str): Mensagem de erro a ser exibida.
        titulo (str): Título da janela do alerta. Padrão é "Erro".
    """
    messagebox.showerror(titulo, mensagem)  # Exibe o alerta de erro

def for_lista_print(lista,tipo=0):
    for i in lista:
        print(i)

def abrir_arquivo(caminho):
    # Define o caminho do arquivo
    file_path = caminho

    # Verifica se o arquivo existe
    if not os.path.exists(file_path):
        # Exibe uma mensagem de erro se o arquivo não for encontrado
        messagebox.showwarning("Erro", "Arquivo Excel não encontrado!")
        return None, None

    # Carrega o arquivo Excel
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Cria uma lista para armazenar as linhas do Excel
    lista_oficial = []
    
    # Itera sobre as linhas da planilha, começando da segunda linha
    for n, i in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
        lista = list(i)
        # Insere o número da linha no início da lista
        lista.insert(0, n + 2)
        lista_oficial.append(lista)
    
    # Retorna a planilha ativa
    return sheet

def carregar_dados(dados,tree):
    dados_para_uso = abrir_arquivo(dados)

    for item in tree.get_children():
        # Deleta cada item
        tree.delete(item)

    dados = list(dados_para_uso.iter_rows(min_row=2, values_only=True))
    for item in reversed(dados):
        tree.insert("", "end", values=item)


def tabela(dados,Freme,linha_de_quantidade_colunas):
    dados_para_uso = abrir_arquivo(dados)

    lista_nomes_de_colunas = []

    for n,t in enumerate(dados_para_uso.iter_rows(values_only=True)):
        if n == linha_de_quantidade_colunas:
            for i in t:
                lista_nomes_de_colunas.append(i)
        break

    tree = ttk.Treeview(Freme, columns=lista_nomes_de_colunas, show="headings")

    for nomes_colunas in lista_nomes_de_colunas:
        tree.heading(nomes_colunas, text=nomes_colunas)
        tree.column(nomes_colunas, width=100)
    
    return tree
    
def carregar_dados_e_tabela_filtro(dados,Freme):
    for item in Freme.get_children():
        Freme.delete(item)
    
    for item in reversed(dados):
        Freme.insert("", "end", values=item)

    return dados

def filtro(app,dados,data_inicial,data_final,busca_por_nome,indice_de_coluna_data,indice_de_coluna_nome):

    print(type(data_inicial), type(data_final))

    tabela = abrir_arquivo(dados)

    lista =[]

    lista_nomes_de_colunas = []

    for n,t in enumerate(tabela.iter_rows(values_only=True)):
        if n == 0:
            for i in t:
                lista_nomes_de_colunas.append(i)

    print(lista_nomes_de_colunas)

    for i in tabela.iter_rows(min_row=2, values_only=True):
        # Verificando se i[2] não é None e é um objeto datetime
        if busca_por_nome in i[indice_de_coluna_nome]:
            if i[indice_de_coluna_data] is not None and isinstance(i[indice_de_coluna_data], datetime):
                data = i[indice_de_coluna_data].date()  # Obtendo a parte da data
                print(data_inicial <= data <= data_final,data_inicial,data,data_final)
                # Verificando se a data está dentro do intervalo
                if data_inicial <= data <= data_final:
                    lista.append(i)
    
    print(lista)
    
    for item in app.get_children():
        app.delete(item)
    
    for i in reversed(lista):
        app.insert("", "end", values=i)

    return lista

def deletar_item(freme_da_tabela,dados):
    # Obtém o item selecionado no Treeview
    wb = load_workbook(dados)
    ws = wb.active

    item_selecionado = freme_da_tabela.selection()


    if item_selecionado:  # Verifica se algum item foi selecionado
        # Deleta o item selecionado
        valores = freme_da_tabela.item(item_selecionado, "values")
        decisao = mostrar_alerta(f"Deseja Realmente excluir estes registro{valores[0],valores[1]}")

        if decisao == True:
            for index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # Compara o valor da célula (primeira coluna) com o valor do Treeview
                if int(row[0]) == int(valores[0]):  # Supondo que o identificador está na primeira coluna
                    ws.delete_rows(index)  # Deleta a linha correspondente
                    break  # Encerra o loop após encontrar a linha
        else:
            print("Operação cancelada pelo usuário.")
    else:
        print("Nenhum item selecionado para deletar.")

    # Salva as alterações no arquivo Excel
    wb.save(dados)


def validar_numeros_decimais(valor):
    try:
        float(valor)  # Verifica se é um número válido
        return True
    except ValueError:
        return valor == ""
    
def deletar_pre_lista(tabela,lista):
   
   if tabela.selection():
        item_selecionado = tabela.selection()
        iten = tabela.item(item_selecionado, "values")

        permisao = mostrar_alerta(iten)

        if permisao:
            for id,item_id in enumerate(tabela.get_children()):  # Itera pelos IDs das linhas
                # valores = tabela.item(item_id, "values")  # Obtém os valores da linha
                for n,x in enumerate(lista):
                    if int(x[0]) == int(iten[0]):
                        lista.pop(n)
                        break

            tabela.delete(item_selecionado)
        return lista
   else:
       messagebox.showwarning("Erro", f"Iten não selecionado")
       return lista
    # except:
    #     messagebox.showwarning("Erro", f"Iten não selecionado")

def carregar_dados_entry(arquivo, indice):
    """
    Carrega dados de uma planilha e retorna uma lista contendo os valores exclusivos
    de uma coluna específica, identificada pelo índice fornecido.
    
    :param arquivo: Caminho para o arquivo (normalmente Excel) que será aberto.
    :param indice: Índice da coluna cujos valores serão extraídos.
    :return: Lista de valores únicos da coluna especificada ou uma mensagem de erro.
    """
    # Abre o arquivo fornecido e carrega os dados
    dados = abrir_arquivo(arquivo)  # Presume que abrir_arquivo é uma função previamente definida.

    # Inicializa a lista que armazenará os valores da coluna especificada
    lista_dados = []
    try:
        # Itera sobre as linhas da planilha, ignorando o cabeçalho (a primeira linha)
        for x in dados.iter_rows(min_row=2, values_only=True):
            # Verifica se o valor da coluna ainda não está na lista; se não, adiciona
            if x[indice] not in lista_dados:
                lista_dados.append(x[indice])
    except:
        # Em caso de erro (como arquivo inválido ou índice fora do alcance), retorna mensagem de erro
        lista_dados = "Dados não Disponível" 

    # Retorna a lista com os dados únicos ou a mensagem de erro
    return lista_dados


def configurar_busca_combobox(combobox, lista_valores):
    """
    Configura o comportamento de busca dinâmica para um Combobox.
    
    :param combobox: O Combobox a ser configurado.
    :param lista_valores: A lista original de valores do Combobox.
    :param delay: Tempo de espera (em milissegundos) após a digitação antes de filtrar.
    """

    delay=600
    filtro_id = None  # Armazena o ID do after para evitar execuções desnecessárias

    def atualizar_lista(event=None):
        nonlocal filtro_id
        if filtro_id:
            combobox.after_cancel(filtro_id)  # Cancela o último after programado
        
        # Aguarda um pouco para processar a entrada
        filtro_id = combobox.after(delay, lambda: filtrar_lista(combobox, lista_valores))

    def filtrar_lista(combobox, lista_valores):
        entrada = combobox.get().lower()
        if not entrada:  # Se a entrada estiver vazia, restaura a lista original
            combobox['values'] = lista_valores
            combobox.event_generate('<Down>')
        else:  # Filtra a lista com base na entrada
            filtrado = [item for item in lista_valores if entrada in str(item).lower()]
            combobox['values'] = filtrado

         # Reabre o Combobox automaticamente (opcional)
            combobox.event_generate('<Down>')
        

    # Adiciona o evento de keyrelease ao Combobox
    combobox.bind("<KeyRelease>", atualizar_lista)

def converter_para_maiusculo(event):
    """Converte o texto do Entry que disparou o evento para maiúsculas."""
    widget = event.widget  # Obtém o widget que disparou o evento
    if isinstance(widget, tk.Entry):  # Garante que seja um Entry
        texto = widget.get()
        widget.delete(0, tk.END)
        widget.insert(0, texto.upper())

def verificar_campos_vazios(campos):
    """Verifica se algum campo está vazio."""
    for widget in campos:
        try:
            if not widget.get().strip():  # Verifica se está vazio ou só tem espaços
                messagebox.showwarning("Campo Vazio", f"Verifique os capos de preenchimento")
                return False
        except:
            if not widget.data_get().strip():  # Verifica se está vazio ou só tem espaços
                messagebox.showwarning("Campo Vazio", f"Verifique os capos de preenchimento")
                return False
    # messagebox.showinfo("Sucesso", "Todos os campos estão preenchidos!")
    return True
                



