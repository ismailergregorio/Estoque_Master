import sys
import os

# Adiciona o diretório do arquivo de funções ao sys.path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from funcoes import abrir_arquivo
from openpyxl import Workbook
import datetime


cadastro = "ajustador_da_tabela_de_dados/001/CADASTRO.xlsx"
dados_de_saida = "001\saida.xlsx"
dados_de_entrada = "001\entrada.xlsx"
dados_de_estoque = "001\Estoque.xlsx"
dados_de_cadastro_nf = "001\Cadastro NF.xlsx"
dados_de_cadastro_fornecedor = "001\Cadastro Fornecedor.xlsx"

def criar_arquivo(nome_arquivo, dados, nome_planilha="Planilha1"):
    workbook = Workbook()
    planilha = workbook.active
    planilha.title = nome_planilha

    # Adiciona os dados linha por linha
    novas_lista = dados
    for i in novas_lista:
        planilha.append(i)

    # Salva o arquivo Excel
    workbook.save(nome_arquivo)
    print(f"Dados salvos com sucesso no arquivo '{nome_arquivo}'.")

def mudar_para_data(arquivo, indice_da_data):
    tabela = abrir_arquivo(arquivo)
    lista = []

    for i in tabela.iter_rows(min_row=2, values_only=True):

        listaa = list(i)

        if type(listaa[indice_da_data]) == int:
            data = datetime.datetime(1900, 1, 1) + datetime.timedelta(days=listaa[indice_da_data] - 2)
            listaa[indice_da_data] = data
            lista.append(listaa)


        elif listaa[indice_da_data] == None:
            print(listaa)
            lista.append(listaa)


        elif type(listaa[indice_da_data]) == str:
            dados_com_hora = listaa[indice_da_data] +  " 00:00:00"
            # Convertendo para datetime
            try:
                data = datetime.datetime.strptime(dados_com_hora, "%d/%m/%Y %H:%M:%S")
            except:
                data = datetime.datetime.strptime(dados_com_hora, "%Y/%m/%d %H:%M:%S")
            listaa[indice_da_data] = data
            lista.append(listaa)

        # print(listaa[indice_da_data])
    
            
    return lista

def codficar(arquivoA,arquivoB):

    tabelaA = abrir_arquivo(arquivoA)
    tabelaB = abrir_arquivo(arquivoB)

    lista_Oficial = []

    for i in tabelaA.iter_rows(min_row=2, values_only=True):
        for x in tabelaB.iter_rows(min_row=2, values_only=True):
            tupla_iten = list(i)
            tupla_itenA = list(i)

            # print(type(tupla_iten),tupla_iten)
            if tupla_iten[0] == x[1]:
                # print(x[0],tupla_iten)
                tupla_iten.insert(0,x[0])
                lista_Oficial.append(tupla_iten)
                
    return lista_Oficial                


tabela = "ajustador_da_tabela_de_dados/001/Cadastro NF.xlsx"


# criar arquivo com o codigo correto
def ajustar_data_e_codificar(tabela,nome):
    dados = codficar(tabela,cadastro)
    criar_arquivo("codificando.xlsx", dados)

    dadosa = mudar_para_data("codificando.xlsx", 2)
    criar_arquivo(nome, dadosa)

    if os.path.exists("codificando.xlsx"):  # Verifica se o arquivo existe
        os.remove("codificando.xlsx")
        print("Arquivo deletado com sucesso.")
    else:
        print("O arquivo não existe.")

def ajustar_data(nome_arquivo,indice):
    dadosa = mudar_para_data("ajustador_da_tabela_de_dados/001/CADASTRO.xlsx", indice)
    criar_arquivo(nome_arquivo, dadosa)

# ajustar_data_e_codificar(tabela,"Cadastro NF.xlsx")

ajustar_data("CADASTRO.xlsx",2)

