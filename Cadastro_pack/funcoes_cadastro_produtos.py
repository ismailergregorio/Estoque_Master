from funcoes import*
import tkinter as tk
from tkinter import messagebox

import openpyxl
import os
from openpyxl.utils.datetime import from_excel
from openpyxl import load_workbook



cadastro = "Base de dados\CADASTRO.xlsx"
dados_de_saida = "Base de dados\saida.xlsx"
dados_de_entrada = "Base de dados\entrada.xlsx"
dados_de_estoque = "Base de dados\Estoque.xlsx"
dados_de_cadastro_nf = "Base de dados\Cadastro NF.xlsx"
dados_de_cadastro_fornecedor = "Base de dados\Cadastro Fornecedor.xlsx"

nomes_dos_setores = ["RH", "ESCRITORIO", "SEGURANÇA", "COSINHA LIMPESA", "COSINHA COMIDA", "TRADE", "OPERACIONAL"]


def abrir_arquivo(caminho):
    # Define o caminho do arquivo
    file_path = caminho

    # Verifica se o arquivo existe
    if not os.path.exists(file_path):
        # Exibe uma mensagem de erro se o arquivo não for encontrado
        messagebox.showwarning("Erro", "Arquivo Excel não encontrado!")
        return None, None

    # Carrega o arquivo Excel
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Cria uma lista para armazenar as linhas do Excel
    lista_oficial = []
    
    # Itera sobre as linhas da planilha, começando da segunda linha
    for n, i in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
        lista = list(i)
        # Insere o número da linha no início da lista
        lista.insert(0, n + 2)
        lista_oficial.append(lista)
    
    # Retorna a planilha ativa
    return sheet

def calcular_estoque_desejavel(event,entry_estoque_minimo="",entry_estoque_desejavel=""):
    try:
        if entry_estoque_minimo.get() == "":
            entry_estoque_desejavel.insert(0, f"{0}")
        else:
            estoque_minimo = float(entry_estoque_minimo.get())
            estoque_desejavel = int(estoque_minimo + (estoque_minimo / 3))
            entry_estoque_desejavel.delete(0, tk.END)
            entry_estoque_desejavel.insert(0, f"{estoque_desejavel:.2f}")
    except ValueError:
        entry_estoque_desejavel.delete(0, tk.END)
        entry_estoque_desejavel.insert(0, "0.00")

def verificar_lista(itens):
    # Verifica se todos os campos da lista foram preenchidos
    for entrada in itens:
        if isinstance(entrada, float):  # Caso seja float, verifica se é maior que zero
            if entrada <= 0:
                messagebox.showwarning("Aviso", "Há valores numéricos inválidos na lista!")
                return False
        elif not entrada or not entrada.strip():  # Verifica se o campo de texto está vazio
            messagebox.showwarning("Aviso", "Todos os campos devem ser preenchidos!")
            return False
    # Se todos os campos forem válidos
    messagebox.showinfo("Sucesso", "Todos os campos foram preenchidos corretamente!")
    return True

def cadastrar_produto(dados,entry_nome_produto,entry_setor,entry_data,entry_estoque_m,entry_estoque_d,entry_obs):

    arquivo = abrir_arquivo(dados)
    maior = 0
    for itens in arquivo.iter_rows(min_row=2,values_only=True):
        if itens[0] > maior:
            maior = itens[0]
        elif entry_nome_produto == itens[1]:
            print(f"Itens")
            break

    codigo_final=maior+1

    lista = [codigo_final,entry_nome_produto,entry_setor,entry_data,float(entry_estoque_m),float(entry_estoque_d),entry_obs]


    workbook = load_workbook(dados)
    sheet = workbook.active  # Seleciona a planilha ativa

    workbook_estoque = load_workbook(dados_de_estoque)
    sheet_estoque = workbook_estoque.active  # Seleciona a planilha ativa

    # Adicionar novos dados em uma nova linha
    novos_dados = lista
    sheet.append(novos_dados)

    novos_dados_estoque = [codigo_final,entry_nome_produto,entry_setor,entry_estoque_m,0,0,0,0]
    sheet_estoque.append(novos_dados_estoque)

    # Salvar o arquivo
    workbook.save(dados)
    workbook_estoque.save(dados_de_estoque)
    print(f"Dados adicionados com sucesso!{lista}")

def deletar_item_cadastro(freme_da_tabela,dados):
    # Obtém o item selecionado no Treeview
    wb = load_workbook(dados)
    ws = wb.active

    wb_estoque = load_workbook(dados_de_estoque)
    ws_estoque = wb_estoque.active

    item_selecionado = freme_da_tabela.selection()


    if item_selecionado:  # Verifica se algum item foi selecionado
        # Deleta o item selecionado
        valores = freme_da_tabela.item(item_selecionado, "values")
        decisao = mostrar_alerta(f"Deseja Realmente excluir estes registro{valores[0],valores[1]}")

        if decisao == True:
            for index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # Compara o valor da célula (primeira coluna) com o valor do Treeview
                if int(row[0]) == int(valores[0]):  # Supondo que o identificador está na primeira coluna
                    ws.delete_rows(index)  # Deleta a linha correspondente
                    break  # Encerra o loop após encontrar a linha
            
            for index, row in enumerate(ws_estoque.iter_rows(min_row=2, values_only=True), start=2):
                # Compara o valor da célula (primeira coluna) com o valor do Treeview
                if int(row[0]) == int(valores[0]):  # Supondo que o identificador está na primeira coluna
                    ws_estoque.delete_rows(index)  # Deleta a linha correspondente
                    break  # Encerra o loop após encontrar a linha
        else:
            print("Operação cancelada pelo usuário.")
    else:
        print("Nenhum item selecionado para deletar.")

    # Salva as alterações no arquivo Excel
    wb_estoque.save(dados_de_estoque)
    wb.save(dados)


