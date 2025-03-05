import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from datetime import datetime
import openpyxl
import os
import re
from tkcalendar import DateEntry
from openpyxl.utils.datetime import from_excel
import time

import datetime

def formata_data(valor_date):
    # Verificar se o valor é None
    if valor_date is None:
        return "Não definido"
    
    # Verificar se o valor é um número de série do Excel
    if isinstance(valor_date, (int, float)):
        python_date = from_excel(valor_date)
        # Formatar a data no formato dd/mm/aaaa
        formatted_date = python_date.strftime('%d/%m/%Y')
    # Verificar se o valor é um datetime
    elif isinstance(valor_date, datetime.datetime):
        formatted_date = valor_date.strftime('%d/%m/%Y')
    else:
        formatted_date = valor_date  # Caso não seja um número de série ou datetime, use o valor original
    
    return formatted_date

def formatar_em_reais(valor):
    try:
        if valor is None:
            return "R$ 0,00"
        return f"R$ {float(valor):,.2f}".replace(",", ".")
    except ValueError:
        return "R$ 0,00"

def formatar_em_float(valor):
    try:
        if valor is None:
            return "Valor não Definido"
        return f"{float(valor):,.2f}"
    except ValueError:
        return "Erro de valor"
    
def carregar_dados():
    file_path = "Base de dados\saida.xlsx"
    if not os.path.exists(file_path):
        messagebox.showwarning("Erro", "Arquivo Excel não encontrado!")
        return
        
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    
    # Adicionar os dados ao Treeview e salvar os itens
    global tree_items
    tree_items = []

    file_path2 = "Base de dados\saidai.xlsx"
    
    # Cria um novo arquivo se ele não existir, caso contrário, abre o existente
    if not os.path.exists(file_path2):
        workbook = openpyxl.Workbook()
        sheet2 = workbook.active
        sheet2.title = "Produtos"
        sheet2.append(["Produto	Setor","Setor","Data","Quant Saida","Motivo Saida","Funcionario","Setor(TRABALHO)","VALOR","OBS"])
    else:
        workbook = openpyxl.load_workbook(file_path2)
        sheet2 = workbook.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        updated_row = list(row)
        
        data = updated_row[2]
        valor_total = updated_row[7]

        # Formatar os valores de data e moeda
        updated_row[2] = formata_data(data)
        updated_row[7] = formatar_em_reais(valor_total)

        # item_id = tree.insert("", "end", values=updated_row)
        # tree_items.append(item_id)
        # file_path = "entrada.xlsx"
    
        # Cria um novo arquivo se ele não existir, caso contrário, abre o existente
        # if not os.path.exists(file_path):
        #     workbook = openpyxl.Workbook()
        #     sheet = workbook.active
        #     sheet.title = "Produtos"
        #     sheet.append(["Produto", "Setor", "Data", "Quantidade","Motivo","Valor Unitario","Valor Total","Observação"])
        # else:
        #     workbook = openpyxl.load_workbook(file_path)
        #     sheet = workbook.active

        # Adiciona os dados na planilha
        sheet2.append(updated_row)
        print(updated_row)
        workbook.save(file_path2)
        # messagebox.showinfo("Sucesso", "Dados salvos com sucesso no Excel!")


tree_items=[]
carregar_dados()